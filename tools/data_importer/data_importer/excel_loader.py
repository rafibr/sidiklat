"""Load training data from Excel workbooks."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook

from .models import ExcelTrainingRow


DEFAULT_COLUMN_MAP: Dict[str, str] = {
    "nama": "nama",
    "nip": "nip",
    "jabatan": "jabatan",
    "unit_kerja": "unit_kerja",
    "nama_pelatihan": "nama_pelatihan",
    "jenis_pelatihan": "jenis_pelatihan",
    "penyelenggara": "penyelenggara",
    "tempat": "tempat",
    "tanggal_mulai": "tanggal_mulai",
    "tanggal_selesai": "tanggal_selesai",
    "jp": "jp",
    "sertifikat_url": "sertifikat_url",
    "keterangan": "keterangan",
}


def _normalize_header(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    return value.strip().lower().replace(" ", "_")


def _parse_date(value) -> Optional[datetime.date]:
    if value in (None, "", "-"):
        return None

    if isinstance(value, datetime):
        return value.date()

    if hasattr(value, "date"):
        return value.date()

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %B %Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(value) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


class ExcelLoader:
    """Loader that converts workbook rows to :class:`ExcelTrainingRow`."""

    def __init__(self, column_map: Optional[Dict[str, str]] = None) -> None:
        self.column_map = column_map or DEFAULT_COLUMN_MAP

    def load(self, path: Path, sheet_name: Optional[str] = None) -> List[ExcelTrainingRow]:
        workbook = load_workbook(filename=path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_index = {
            _normalize_header(header): idx for idx, header in enumerate(header_row) if header is not None
        }

        rows: List[ExcelTrainingRow] = []
        for raw_row in sheet.iter_rows(min_row=2, values_only=True):
            data = {}
            for field, header_name in self.column_map.items():
                key = _normalize_header(header_name)
                if key not in header_index:
                    continue
                data[field] = raw_row[header_index[key]]

            if not data.get("nama") and not data.get("nama_pelatihan"):
                continue

            rows.append(
                ExcelTrainingRow(
                    nama=str(data.get("nama") or "").strip(),
                    nip=str(data.get("nip") or "").strip() or None,
                    jabatan=(str(data.get("jabatan")).strip() if data.get("jabatan") else None),
                    unit_kerja=(str(data.get("unit_kerja")).strip() if data.get("unit_kerja") else None),
                    nama_pelatihan=str(data.get("nama_pelatihan") or "").strip(),
                    jenis_pelatihan=(str(data.get("jenis_pelatihan")).strip() if data.get("jenis_pelatihan") else None),
                    penyelenggara=(str(data.get("penyelenggara")).strip() if data.get("penyelenggara") else None),
                    tempat=(str(data.get("tempat")).strip() if data.get("tempat") else None),
                    tanggal_mulai=_parse_date(data.get("tanggal_mulai")),
                    tanggal_selesai=_parse_date(data.get("tanggal_selesai")),
                    jp=_parse_int(data.get("jp")),
                    sertifikat_url=(str(data.get("sertifikat_url")).strip() if data.get("sertifikat_url") else None),
                    keterangan=(str(data.get("keterangan")).strip() if data.get("keterangan") else None),
                )
            )
        return rows

    def load_iter(self, path: Path, sheet_name: Optional[str] = None) -> Iterable[ExcelTrainingRow]:
        for row in self.load(path, sheet_name):
            yield row
